import openpyxl
import openpyxl.utils
import pandas as pd
from pathlib import Path

from providers.base_provider import BaseProvider

# Fallback when a provider config's extraction block omits header_row: assume the header is on
# the first row. Unlike header_row, data_start_row's fallback (header_row + 1) has no separate
# constant — it's inherently derived from whatever header_row resolves to.
DEFAULT_HEADER_ROW_1IDX = 1


class ExcelProvider(BaseProvider):
    """Provider that extracts data from Excel files using column-index-based config.

    Works for any provider whose config declares an 'extraction' block with
    header_row, data_start_row, column mappings (column_index + source_header),
    and a data-end detection strategy.
    """

    def extract(self, file_path) -> pd.DataFrame:
        """Read the Excel file and return a DataFrame with standardized columns."""
        file_path = Path(file_path)
        self.logger.info(f'[{self.name}] Extracting data from: {file_path.name}')

        extraction_cfg = self.config.get_value('extraction')
        if not extraction_cfg:
            raise ValueError(f'[{self.name}] Missing "extraction" section in provider config.')

        sheet_name = extraction_cfg.get('sheet_name')  # None → active sheet

        header_row_cfg = extraction_cfg.get('header_row')
        if header_row_cfg is None:
            header_row_1idx = DEFAULT_HEADER_ROW_1IDX
            self.logger.warning(
                f'Provider "{self.name}" config is missing "header_row" in its extraction block; '
                f'defaulting to row {DEFAULT_HEADER_ROW_1IDX}.'
            )
        else:
            header_row_1idx = int(header_row_cfg)

        data_start_row_cfg = extraction_cfg.get('data_start_row')
        if data_start_row_cfg is None:
            data_start_row_1idx = header_row_1idx + 1
            self.logger.warning(
                f'Provider "{self.name}" config is missing "data_start_row" in its extraction block; '
                f'defaulting to header_row + 1 = {data_start_row_1idx}.'
            )
        else:
            data_start_row_1idx = int(data_start_row_cfg)

        data_end_strategy = extraction_cfg.get('data_end_strategy', 'first_empty')
        data_end_anchor = extraction_cfg.get('data_end_anchor_field', None)
        columns_cfg: dict = extraction_cfg.get('columns', {})

        if not columns_cfg:
            raise ValueError(f'[{self.name}] No columns defined in provider config extraction.columns.')

        # Build a mapping: output_field -> 0-based column index
        col_map = {}
        for field, cfg in columns_cfg.items():
            col_1idx = int(cfg.get('column_index'))
            col_map[field] = col_1idx - 1  # convert to 0-based

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            self.logger.info(f'[{self.name}] Using sheet: "{ws.title}"')

            # Validate headers against what the config expects
            header_row_values = next(
                ws.iter_rows(min_row=header_row_1idx, max_row=header_row_1idx, values_only=True)
            )
            self._validate_headers(header_row_values, columns_cfg, col_map)

            # Extract data rows
            records = []
            anchor_col_idx = col_map.get(data_end_anchor) if data_end_anchor else None

            for row_values in ws.iter_rows(min_row=data_start_row_1idx, values_only=True):
                # End-of-data detection
                if data_end_strategy == 'first_empty' and anchor_col_idx is not None:
                    anchor_val = row_values[anchor_col_idx] if anchor_col_idx < len(row_values) else None
                    if anchor_val is None or str(anchor_val).strip() == '':
                        break

                record = {}
                for field, col_0idx in col_map.items():
                    record[field] = row_values[col_0idx] if col_0idx < len(row_values) else None
                records.append(record)
        finally:
            wb.close()

        df = pd.DataFrame(records, columns=list(col_map.keys()))
        self.logger.info(f'[{self.name}] Extracted {len(df)} data rows.')
        return df

    def _validate_headers(self, header_row_values: tuple, columns_cfg: dict, col_map: dict):
        """Log warnings when a column's actual header does not match the expected value in config."""
        for field, cfg in columns_cfg.items():
            expected_header = cfg.get('source_header', '').strip()
            col_0idx = col_map[field]
            actual_header = (
                str(header_row_values[col_0idx]).strip()
                if col_0idx < len(header_row_values) and header_row_values[col_0idx] is not None
                else ''
            )
            col_1idx = col_0idx + 1
            col_letter = openpyxl.utils.get_column_letter(col_1idx)
            if expected_header and actual_header.lower() != expected_header.lower():
                self.logger.warning(
                    f'[{self.name}] Header mismatch for field "{field}": '
                    f'expected "{expected_header}", found "{actual_header}" '
                    f'at column {col_letter} (index {col_1idx}).'
                )
            else:
                self.logger.debug(
                    f'[{self.name}] Header OK for field "{field}": "{actual_header}" '
                    f'at column {col_letter} (index {col_1idx}).'
                )
